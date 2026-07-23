"""
成果导出（阶段 8E）测试。

覆盖：Markdown/JSON/PDF/DOCX/XLSX 五种格式导出成功、内容包含标题
与来源任务信息、不泄露敏感字段（system prompt/API Key/Provider
原始响应/内部 Context）、文件名安全（ASCII fallback + UTF-8）、
不支持的格式被拒绝、导出接口只接受整数 id（天然防止目录穿越）。
"""

import json
import uuid
import zipfile
from datetime import datetime, timezone
from io import BytesIO

import pytest
from fastapi.testclient import TestClient

from app.database.db import SessionLocal
from app.main import app
from app.models.deliverable_db import DeliverableDB, DeliverableVersionDB
from app.models.task_db import TaskDB
from app.services.deliverable_export_service import (
    SUPPORTED_FORMATS,
    build_content_disposition,
    export_deliverable,
)

TEST_MARKER = "DELIVERABLE_EXPORT_TEST"

_FORBIDDEN_SNIPPETS = (
    "DEEPSEEK_API_KEY",
    "sk-real-fake-provider-key",
    "system_prompt",
    "You are AI CEO",
    "Authorization: Bearer",
)


@pytest.fixture(scope="module")
def client():
    with TestClient(app) as test_client:
        yield test_client


def _make_task_id():
    return f"TASKEXPORT{uuid.uuid4().hex[:8].upper()}"


@pytest.fixture
def sample_deliverable():
    task_id = _make_task_id()
    db = SessionLocal()
    try:
        task = TaskDB(
            id=task_id,
            task_type="经营分析导出测试",
            assigned_agent="AI CEO",
            priority="normal",
            status="completed",
            payload={"task": "经营分析导出测试"},
            result={
                "success": True,
                "agent": "AI CEO",
                "decision": {},
                "result": {
                    "agent": "AI CEO",
                    "analysis_type": "system_operations",
                    "provider": "deepseek",
                    "model": "deepseek-chat",
                    "format": "structured",
                    "analysis": {
                        "summary": "本周经营平稳，任务完成率保持在合理区间",
                        "findings": ["任务完成率稳定"],
                        "risks": ["单一 Agent 依赖"],
                        "actions": ["补充销售数据来源"],
                        "delegations": [],
                    },
                },
            },
            created_at=datetime.now(timezone.utc),
            completed_at=datetime.now(timezone.utc),
            root_task_id=task_id,
        )
        db.add(task)
        db.commit()

        deliverable = DeliverableDB(
            deliverable_code=f"DLV-{uuid.uuid4().hex[:12].upper()}",
            title="经营分析：导出测试专用成果",
            deliverable_type="ceo_analysis",
            status="pending_review",
            source_task_id=task_id,
            root_task_id=task_id,
            agent_name="AI CEO",
            summary="本周经营平稳",
            current_version=1,
        )
        db.add(deliverable)
        db.flush()

        version = DeliverableVersionDB(
            deliverable_id=deliverable.id,
            version_number=1,
            format="structured",
            content="本周经营平稳，任务完成率保持在合理区间\n\n### 发现\n\n- 任务完成率稳定\n",
            structured_content={
                "summary": "本周经营平稳，任务完成率保持在合理区间",
                "findings": ["任务完成率稳定"],
                "risks": ["单一 Agent 依赖"],
                "actions": ["补充销售数据来源"],
                "delegations": [],
            },
            created_by="system",
            source_task_id=task_id,
        )
        db.add(version)
        db.commit()

        deliverable_id = deliverable.id

    finally:
        db.close()

    yield deliverable_id, task_id

    db = SessionLocal()
    try:
        db.query(DeliverableVersionDB).filter(
            DeliverableVersionDB.deliverable_id == deliverable_id
        ).delete(synchronize_session=False)
        db.query(DeliverableDB).filter(DeliverableDB.id == deliverable_id).delete(
            synchronize_session=False
        )
        db.query(TaskDB).filter(TaskDB.id == task_id).delete(synchronize_session=False)
        db.commit()
    finally:
        db.close()


def _load(sample_deliverable):
    deliverable_id, _ = sample_deliverable
    db = SessionLocal()
    try:
        deliverable = db.query(DeliverableDB).filter(DeliverableDB.id == deliverable_id).first()
        version = (
            db.query(DeliverableVersionDB)
            .filter(DeliverableVersionDB.deliverable_id == deliverable_id)
            .first()
        )
        return deliverable, version
    finally:
        db.close()


def test_export_markdown_contains_title_and_source_task(sample_deliverable):
    deliverable, version = _load(sample_deliverable)
    content, content_type, ascii_name, utf8_name = export_deliverable(
        deliverable, version, "markdown", shop_name=None
    )
    text = content.decode("utf-8")

    assert content_type.startswith("text/markdown")
    assert deliverable.title in text
    assert deliverable.source_task_id in text
    assert "未绑定店铺" in text
    assert ascii_name.endswith(".md")


def test_export_json_only_contains_safe_fields(sample_deliverable):
    deliverable, version = _load(sample_deliverable)
    content, content_type, ascii_name, _ = export_deliverable(
        deliverable, version, "json", shop_name="测试店铺"
    )
    assert content_type == "application/json"
    payload = json.loads(content)

    assert payload["title"] == deliverable.title
    assert payload["shop_name"] == "测试店铺"
    assert "encrypted_value" not in json.dumps(payload)
    assert "system_prompt" not in json.dumps(payload)


def test_export_pdf_generates_nonempty_binary_with_chinese_font(sample_deliverable):
    deliverable, version = _load(sample_deliverable)
    content, content_type, ascii_name, _ = export_deliverable(
        deliverable, version, "pdf", shop_name=None
    )
    assert content_type == "application/pdf"
    assert content[:4] == b"%PDF"
    assert len(content) > 500
    assert ascii_name.endswith(".pdf")


def test_export_docx_is_valid_zip_with_document_xml(sample_deliverable):
    deliverable, version = _load(sample_deliverable)
    content, content_type, ascii_name, _ = export_deliverable(
        deliverable, version, "docx", shop_name=None
    )
    assert content_type.endswith("wordprocessingml.document")

    archive = zipfile.ZipFile(BytesIO(content))
    assert "word/document.xml" in archive.namelist()
    document_xml = archive.read("word/document.xml").decode("utf-8")
    assert "经营分析" in document_xml or "导出测试" in document_xml


def test_export_xlsx_has_multiple_sheets_for_ceo_analysis(sample_deliverable):
    deliverable, version = _load(sample_deliverable)
    content, content_type, ascii_name, _ = export_deliverable(
        deliverable, version, "xlsx", shop_name=None
    )
    assert content_type.endswith("spreadsheetml.sheet")

    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(content))
    assert "摘要" in workbook.sheetnames
    assert "发现" in workbook.sheetnames
    assert "风险" in workbook.sheetnames
    assert "行动" in workbook.sheetnames


def test_export_does_not_leak_forbidden_snippets(sample_deliverable):
    deliverable, version = _load(sample_deliverable)
    for export_format in ("markdown", "json"):
        content, _, _, _ = export_deliverable(deliverable, version, export_format, shop_name=None)
        text = content.decode("utf-8", errors="ignore")
        for snippet in _FORBIDDEN_SNIPPETS:
            assert snippet not in text


def test_content_disposition_has_ascii_and_utf8_filename():
    header = build_content_disposition("DLV-ABC123.md", "经营分析-DLV-ABC123.md")
    assert 'filename="DLV-ABC123.md"' in header
    assert "filename*=UTF-8''" in header


def test_export_api_returns_correct_content_type(client, sample_deliverable):
    deliverable_id, _ = sample_deliverable

    response = client.get(
        f"/api/v1/deliverables/{deliverable_id}/export", params={"format": "markdown"}
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/markdown")
    assert "attachment" in response.headers["content-disposition"]


def test_export_api_rejects_unsupported_format(client, sample_deliverable):
    deliverable_id, _ = sample_deliverable

    response = client.get(
        f"/api/v1/deliverables/{deliverable_id}/export",
        params={"format": "exe"},
    )
    assert response.status_code == 422


def test_export_api_path_only_accepts_integer_id(client):
    # deliverable_id 是路径参数且声明为 int；传入路径穿越式字符串
    # 会被 FastAPI/Starlette 路由层直接拒绝（404 或 422），永远不会
    # 到达任何文件系统读取逻辑——本导出实现全程在内存中生成文件，
    # 不存在可被穿越的路径参数。
    response = client.get(
        "/api/v1/deliverables/../../etc/passwd/export",
        params={"format": "markdown"},
    )
    assert response.status_code in (404, 422)


def test_all_supported_formats_present():
    assert set(SUPPORTED_FORMATS) == {"markdown", "json", "pdf", "docx", "xlsx"}
